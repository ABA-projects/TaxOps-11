"""Generador Excel Formulario 210 DIAN — layout horizontal (multi-casilla por fila).

Columnas A-F:  A(cas#,w6)  B(desc,w30)  C(valor,w16)  |  D(cas#,w6)  E(desc,w30)  F(valor,w16)
Casillas totales: span completo A(cas#) B:E(desc merged) F(valor).
Orientación landscape, colores paleta DIAN oficial.
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── Paleta DIAN ───────────────────────────────────────────────────────────────
_DARK = "1F3864"   # azul oscuro — títulos principales
_MED = "2E75B6"   # azul medio — encabezados de sección
_LIGHT = "BDD7EE"   # azul claro — totales sección
_VLIGHT = "DEEAF1"   # azul muy claro — filas alternas
_WHITE = "FFFFFF"
_YELLOW = "FFFF00"   # BORRADOR banner
_RED = "C00000"   # texto rojo advertencia
_BG_RED = "FFE6E6"   # fondo saldo a pagar
_BG_GREEN = "E2EFDA"   # fondo saldo a favor

UVT_2025 = 49_799.0
LIMIT_UVT_40 = 1_340 * UVT_2025


# ── Helpers de estilos ────────────────────────────────────────────────────────

def _thin() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _write_val(c, value: Any, bold: bool = False, bg: str = _WHITE) -> None:
    """Escribe número formateado o texto/dash en celda de valor."""
    try:
        num = float(value) if value not in (None, "", "—") else None
    except (ValueError, TypeError):
        num = None

    if num is not None:
        c.value = num
        c.number_format = "#,##0"
    elif value not in (None, "", "—"):
        c.value = str(value)
    else:
        c.value = "—"

    c.font = Font(name="Calibri", size=9, bold=bold, color=_DARK)
    c.fill = _fill(bg)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = _thin()


# ── Constructores de filas ────────────────────────────────────────────────────

def _hdr_full(ws, row: int, text: str, color: str,
              text_color: str = "FFFFFF", size: int = 10,
              height: float = 18, indent: int = 1) -> None:
    """Encabezado de sección que ocupa todo el ancho (A:F)."""
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row=row, column=1)
    c.value = text
    c.font = Font(name="Calibri", size=size, bold=True, color=text_color)
    c.fill = _fill(color)
    c.alignment = Alignment(horizontal="left", vertical="center",
                            indent=indent, wrap_text=True)
    c.border = _thin()
    ws.row_dimensions[row].height = height


def _single(ws, row: int, cas: int | str, desc: str, value: Any,
            bg: str = _VLIGHT, bold_val: bool = False, height: float = 15) -> None:
    """Casilla total: A(#) + B:E(desc merged) + F(valor) — fila ancha."""
    c = ws.cell(row=row, column=1, value=str(cas) if cas else "")
    c.font = Font(name="Calibri", size=9, bold=bool(cas), color=_DARK)
    c.fill = _fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _thin()

    ws.merge_cells(f"B{row}:E{row}")
    c = ws.cell(row=row, column=2, value=desc)
    c.font = Font(name="Calibri", size=9, bold=bold_val, color=_DARK)
    c.fill = _fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center",
                            indent=1, wrap_text=True)
    c.border = _thin()

    _write_val(ws.cell(row=row, column=6), value, bold=bold_val, bg=bg)
    ws.row_dimensions[row].height = height


def _pair(ws, row: int,
          l_cas: int | str, l_desc: str, l_val: Any,
          r_cas: int | str, r_desc: str, r_val: Any,
          bg_l: str = _WHITE, bg_r: str = _VLIGHT,
          height: float = 15) -> None:
    """Dos casillas lado a lado: izq (A,B,C) — der (D,E,F)."""
    for col_start, cas, desc, val, bg in [
        (1, l_cas, l_desc, l_val, bg_l),
        (4, r_cas, r_desc, r_val, bg_r),
    ]:
        c = ws.cell(row=row, column=col_start, value=str(cas) if cas else "")
        c.font = Font(name="Calibri", size=9, bold=bool(cas), color=_DARK)
        c.fill = _fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _thin()

        c = ws.cell(row=row, column=col_start + 1, value=desc)
        c.font = Font(name="Calibri", size=9, color=_DARK)
        c.fill = _fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                indent=1, wrap_text=True)
        c.border = _thin()

        _write_val(ws.cell(row=row, column=col_start + 2), val, bg=bg)

    ws.row_dimensions[row].height = height


def _info(ws, row: int, label: str, value: Any,
          bg: str = _WHITE, height: float = 14) -> None:
    """Fila de dato del declarante: A:C(label) D:F(valor)."""
    ws.merge_cells(f"A{row}:C{row}")
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(name="Calibri", size=9, bold=True, color=_DARK)
    c.fill = _fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = _thin()

    ws.merge_cells(f"D{row}:F{row}")
    c = ws.cell(row=row, column=4,
                value=str(value) if value not in (None, "") else "—")
    c.font = Font(name="Calibri", size=9, color=_DARK)
    c.fill = _fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = _thin()
    ws.row_dimensions[row].height = height


def _blank(ws, row: int) -> None:
    for col in range(1, 7):
        ws.cell(row=row, column=col).fill = _fill(_WHITE)
    ws.row_dimensions[row].height = 5


# ── Cálculos intermedios ──────────────────────────────────────────────────────

def _v(d: dict, key: str, default: float = 0.0) -> float:
    return float(d.get(key) or default)


def _intermedios(decl: dict) -> dict:
    """Calcula casillas intermedias no almacenadas directamente en la DB."""
    ing_lab = _v(decl, "ingresos_laborales")
    pen = _v(decl, "aportes_pension")
    afc = _v(decl, "afc_fvp")
    int_viv = _v(decl, "intereses_vivienda")
    med = _v(decl, "medicina_prepagada")
    dep = int(_v(decl, "dependientes"))
    cap = _v(decl, "rentas_capital")
    no_lab = _v(decl, "rentas_no_laborales")
    div = _v(decl, "dividendos")
    goc = _v(decl, "ganancias_ocasionales")
    pat_bruto = _v(decl, "patrimonio_bruto")
    pasivos = _v(decl, "pasivos")

    c34 = max(0.0, ing_lab - pen)
    c36 = min(round(ing_lab * 0.25, 2), 240 * 12 * UVT_2025)
    c37 = afc + c36

    dep_uvt = min(dep * 32 * UVT_2025, ing_lab * 0.10)
    c39 = med + dep_uvt
    c40 = int_viv + c39

    ingresos_netos = max(0.0, ing_lab - pen + cap + no_lab)
    c41 = round(min(c37 + c40, ingresos_netos * 0.40, LIMIT_UVT_40), 2)
    c42 = max(0.0, c34 - c41)

    c91 = c42 + max(0.0, cap) + max(0.0, no_lab)
    c97 = _v(decl, "renta_gravable") or c91
    c115 = max(0.0, goc)

    imp_cargo = _v(decl, "impuesto_cargo")
    detalle = decl.get("detalle_calculo") or decl.get("datos_calculados") or {}
    if isinstance(detalle, dict):
        c116 = float(detalle.get("impuesto_cedula_general",
                                 detalle.get("consolidado", {}).get("impuesto_cedula_general", imp_cargo)))
    else:
        c116 = imp_cargo
    c127 = _v(detalle, "impuesto_ocasional") if isinstance(detalle, dict) else 0.0
    c129 = imp_cargo

    ret = _v(decl, "retenciones")
    c134 = _v(decl, "saldo_pagar")
    c137 = _v(decl, "saldo_favor")

    return {
        "c29": pat_bruto, "c30": pasivos,
        "c31": max(0.0, pat_bruto - pasivos),
        "c32": ing_lab, "c33": pen, "c34": c34,
        "c35": afc, "c36": c36, "c37": c37,
        "c38": int_viv, "c39": c39, "c40": c40,
        "c41": c41, "c42": c42,
        "c58": cap, "c74": no_lab,
        "c91": c91, "c97": c97,
        "c107": div, "c115": c115,
        "c116": c116, "c127": c127, "c129": c129,
        "c132": ret, "c134": c134, "c136": c134,
        "c137": c137,
    }


# ── Generador principal ───────────────────────────────────────────────────────

def generar_excel(declaracion: dict, contribuyente: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "F210"

    # Landscape + anchos de columna
    ws.page_setup.orientation = "landscape"
    ws.column_dimensions["A"].width = 6    # cas# izquierda
    ws.column_dimensions["B"].width = 30   # desc izquierda
    ws.column_dimensions["C"].width = 16   # valor izquierda
    ws.column_dimensions["D"].width = 6    # cas# derecha
    ws.column_dimensions["E"].width = 30   # desc derecha
    ws.column_dimensions["F"].width = 16   # valor derecha

    v = _intermedios(declaracion)
    año = declaracion.get("año_gravable", 2025)
    est = (declaracion.get("estado") or "borrador").upper()
    nombre = (contribuyente.get("nombre_completo") or "").upper()
    nit = contribuyente.get("numero_doc", "")
    ciudad = contribuyente.get("ciudad", "")

    r = 1

    # ── Títulos ───────────────────────────────────────────────────────────────
    _hdr_full(ws, r,
              "DECLARACIÓN DE RENTA Y COMPLEMENTARIO"
              " — PERSONAS NATURALES Y ASIMILADAS RESIDENTES",
              _DARK, size=12, height=24)
    r += 1
    _hdr_full(ws, r,
              f"FORMULARIO 210  —  AÑO GRAVABLE {año}  —  UVT: $ {UVT_2025:,.0f}",
              _DARK, size=11, height=20)
    r += 1

    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(row=r, column=1,
                value="⚠  BORRADOR — NO VÁLIDO COMO DECLARACIÓN ANTE LA DIAN")
    c.font = Font(name="Calibri", size=10, bold=True, color=_RED)
    c.fill = _fill(_YELLOW)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _thin()
    ws.row_dimensions[r].height = 18
    r += 1

    _blank(ws, r)
    r += 1

    # ── Datos del declarante ──────────────────────────────────────────────────
    _hdr_full(ws, r, "DATOS DEL DECLARANTE", _MED, height=16)
    r += 1
    _info(ws, r, "Nombre / Razón Social", nombre, _VLIGHT)
    r += 1
    _info(ws, r, "NIT / Cédula", nit, _WHITE)
    r += 1
    _info(ws, r, "Ciudad", ciudad, _VLIGHT)
    r += 1
    _info(ws, r, "Estado declaración", est, _WHITE)
    r += 1
    _info(ws, r, "Período gravable", año, _VLIGHT)
    r += 1
    _blank(ws, r)
    r += 1

    # ── PATRIMONIO ────────────────────────────────────────────────────────────
    _hdr_full(ws, r, "PATRIMONIO", _MED)
    r += 1
    _pair(ws, r,
          29, "Patrimonio bruto", v["c29"],
          30, "Deudas", v["c30"])
    r += 1
    _single(ws, r, 31, "Patrimonio líquido  (29 − 30)", v["c31"],
            bg=_LIGHT, bold_val=True)
    r += 1
    _blank(ws, r)
    r += 1

    # ── CÉDULA GENERAL ────────────────────────────────────────────────────────
    _hdr_full(ws, r, "CÉDULA GENERAL", _DARK, height=16)
    r += 1

    # Rentas de trabajo (relación laboral)
    _hdr_full(ws, r, "  RENTAS DE TRABAJO  (Art. 103 ET)", _MED, height=15)
    r += 1
    _pair(ws, r,
          32, "Ingresos brutos rentas de trabajo", v["c32"],
          33, "Ingresos no constitutivos de renta — aportes pensión", v["c33"])
    r += 1
    _single(ws, r, 34, "Renta líquida rentas de trabajo  (32 − 33)",
            v["c34"], bg=_VLIGHT, bold_val=True)
    r += 1
    _pair(ws, r,
          35, "Rentas exentas — AFC / FVP / AVC", v["c35"],
          36, "Rentas exentas — 25%  núm. 10  Art. 206 ET", v["c36"])
    r += 1
    _single(ws, r, 37, "Total rentas exentas rentas de trabajo  (35 + 36)",
            v["c37"], bg=_VLIGHT)
    r += 1
    _pair(ws, r,
          38, "Deducciones — Intereses crédito de vivienda", v["c38"],
          39, "Deducciones — Salud, dependientes y otras", v["c39"])
    r += 1
    _single(ws, r, 40, "Total deducciones imputables  (38 + 39)",
            v["c40"], bg=_VLIGHT)
    r += 1
    _single(ws, r, 41,
            "Rentas exentas y deducciones imputables LIMITADAS — 40% / 1.340 UVT  (Art. 336 ET)",
            v["c41"], bg=_LIGHT, height=18)
    r += 1
    _single(ws, r, 42, "Renta líquida ordinaria rentas de trabajo  (34 − 41)",
            v["c42"], bg=_LIGHT, bold_val=True)
    r += 1
    _blank(ws, r)
    r += 1

    # Rentas de trabajo — no relación laboral (43-53)
    _hdr_full(ws, r, "  RENTAS DE TRABAJO — NO RELACIÓN LABORAL  (Art. 103 ET)",
              _MED, height=15)
    r += 1
    _pair(ws, r, 43, "Ingresos brutos", 0,
          44, "Ingresos no constitutivos de renta", 0)
    r += 1
    _single(ws, r, 53, "Renta líquida ordinaria  (46 − 51 − 52)", 0,
            bg=_VLIGHT)
    r += 1
    _blank(ws, r)
    r += 1

    # Rentas de capital (58-73)
    _hdr_full(ws, r, "  RENTAS DE CAPITAL  (Art. 58 ET)", _MED, height=15)
    r += 1
    _pair(ws, r,
          58, "Ingresos brutos rentas de capital", v["c58"],
          59, "Ingresos no constitutivos de renta", 0)
    r += 1
    _pair(ws, r,
          60, "Costos y deducciones procedentes", 0,
          61, "Renta líquida  (58 − 59 − 60)", max(0.0, v["c58"]))
    r += 1
    _pair(ws, r,
          65, "Total rentas exentas", 0,
          68, "Total deducciones imputables", 0)
    r += 1
    _single(ws, r, 73, "Renta líquida ordinaria rentas de capital",
            v["c58"], bg=_VLIGHT, bold_val=True)
    r += 1
    _blank(ws, r)
    r += 1

    # Rentas no laborales (74-90)
    _hdr_full(ws, r, "  RENTAS NO LABORALES  (Art. 74 ET)", _MED, height=15)
    r += 1
    _pair(ws, r,
          74, "Ingresos brutos rentas no laborales", v["c74"],
          75, "Devoluciones, rebajas y descuentos", 0)
    r += 1
    _pair(ws, r,
          76, "Ingresos no constitutivos de renta", 0,
          77, "Costos y deducciones procedentes", 0)
    r += 1
    _single(ws, r, 78, "Renta líquida  (74 − 75 − 76 − 77)",
            max(0.0, v["c74"]), bg=_VLIGHT)
    r += 1
    _pair(ws, r,
          82, "Total rentas exentas", 0,
          85, "Total deducciones imputables", 0)
    r += 1
    _single(ws, r, 90, "Renta líquida ordinaria rentas no laborales",
            v["c74"], bg=_VLIGHT, bold_val=True)
    r += 1
    _blank(ws, r)
    r += 1

    # Consolidado cédula general (91-97)
    _hdr_full(ws, r, "  CONSOLIDADO CÉDULA GENERAL", _MED, height=15)
    r += 1
    _single(ws, r, 91,
            "Renta líquida cédula general  (42 + 53 + 57 + 73 + 90)",
            v["c91"], bg=_VLIGHT, bold_val=True)
    r += 1
    _pair(ws, r,
          93, "Renta líquida ordinaria  (91 − 92)", v["c91"],
          94, "Compensaciones por pérdidas año 2018 y ant.", 0)
    r += 1
    _pair(ws, r,
          95, "Compensaciones por exceso renta presuntiva", 0,
          96, "Rentas gravables", 0)
    r += 1
    _single(ws, r, 97,
            "Renta líquida gravable cédula general  (93 + 96 − 94 − 95)",
            v["c97"], bg=_LIGHT, bold_val=True)
    r += 1
    _blank(ws, r)
    r += 1

    # ── CÉDULA DE PENSIONES (98-103) ──────────────────────────────────────────
    _hdr_full(ws, r, "CÉDULA DE PENSIONES  (Art. 337 ET)", _DARK, height=16)
    r += 1
    _pair(ws, r,
          99, "Ingresos brutos por rentas de pensiones", 0,
          100, "Ingresos no constitutivos de renta", 0)
    r += 1
    _pair(ws, r,
          101, "Renta líquida  (99 − 100)", 0,
          102, "Rentas exentas de pensiones", 0)
    r += 1
    _single(ws, r, 103, "Renta líquida gravable cédula de pensiones  (101 − 102)",
            0, bg=_VLIGHT, bold_val=True)
    r += 1
    _blank(ws, r)
    r += 1

    # ── CÉDULA DE DIVIDENDOS (104-110) ───────────────────────────────────────
    _hdr_full(ws, r,
              "CÉDULA DE DIVIDENDOS Y PARTICIPACIONES  (Art. 242 ET)",
              _DARK, height=16)
    r += 1
    _pair(ws, r,
          104, "Dividendos y part. año 2016 y anteriores", 0,
          105, "Ingresos no constitutivos de renta", 0)
    r += 1
    _pair(ws, r,
          106, "Renta líquida ordinaria 2016 y ant.  (104−105)", 0,
          107, "1a. Subcédula 2017+ — núm. 3 Art. 49 ET", v["c107"])
    r += 1
    _pair(ws, r,
          108, "2a. Subcédula 2017+ — par. 2° Art. 49 ET", 0,
          109, "Dividendos y part. recibidos del exterior", 0)
    r += 1
    _blank(ws, r)
    r += 1

    # ── GANANCIAS OCASIONALES (112-115) ──────────────────────────────────────
    _hdr_full(ws, r, "GANANCIAS OCASIONALES  (Art. 299–316 ET)", _DARK, height=16)
    r += 1
    _pair(ws, r,
          112, "Ingresos por ganancias ocasionales del país y exterior", v["c115"],
          113, "Costos por ganancias ocasionales", 0)
    r += 1
    _pair(ws, r,
          114, "Ganancias ocasionales no gravadas y exentas", 0,
          115, "Ganancias ocasionales gravables  (112−113−114)", v["c115"])
    r += 1
    _blank(ws, r)
    r += 1

    # ── LIQUIDACIÓN PRIVADA ───────────────────────────────────────────────────
    _hdr_full(ws, r, "LIQUIDACIÓN PRIVADA", _DARK, height=16)
    r += 1
    _single(ws, r, 116, "Impuesto sobre renta cédula general  (Art. 241 ET)",
            v["c116"], bg=_VLIGHT, bold_val=True)
    r += 1
    _pair(ws, r,
          119, "Impuesto cédula dividendos 2016  (base cas. 106)", 0,
          120, "Impuesto cédula dividendos 2017 y siguientes", 0)
    r += 1
    _single(ws, r, 121,
            "Impuesto a cargo antes de descuentos  (116 + 119 + 120)",
            v["c116"], bg=_VLIGHT)
    r += 1
    _pair(ws, r,
          122, "Descuentos tributarios — impuestos pagados exterior", 0,
          123, "Descuentos tributarios — donaciones", 0)
    r += 1
    _pair(ws, r,
          124, "Descuentos tributarios — dividendos y participaciones", 0,
          125, "Total descuentos tributarios  (122 + 123 + 124)", 0)
    r += 1
    _single(ws, r, 126, "Impuesto neto de renta  (121 − 125)",
            v["c116"], bg=_VLIGHT)
    r += 1
    _pair(ws, r,
          127, "Impuesto de ganancias ocasionales", v["c127"],
          128, "Descuento impuestos pagados exterior — gan. ocasionales", 0)
    r += 1
    _single(ws, r, 129, "Total impuesto a cargo  (126 + 127 − 128)",
            v["c129"], bg=_LIGHT, bold_val=True)
    r += 1
    _blank(ws, r)
    r += 1

    # Anticipo y retenciones
    _hdr_full(ws, r, "  ANTICIPO Y RETENCIONES", _MED, height=15)
    r += 1
    _pair(ws, r,
          130, "Anticipo renta liquidado año gravable anterior", 0,
          131, "Saldo a favor año gravable anterior", 0)
    r += 1
    _pair(ws, r,
          132, "Retenciones año gravable a declarar", v["c132"],
          133, "Anticipo renta para el año gravable siguiente", 0)
    r += 1
    _blank(ws, r)
    r += 1

    # Saldo a pagar / a favor
    _hdr_full(ws, r, "  SALDO A PAGAR / SALDO A FAVOR", _MED, height=15)
    r += 1
    _pair(ws, r,
          134, "Saldo a pagar por impuesto  (129+133−130−131−132)", v["c134"],
          135, "Sanciones", 0)
    r += 1
    _single(ws, r, 136,
            "Total saldo a pagar  (129 + 133 + 135 − 130 − 131 − 132)", v["c136"],
            bg=_BG_RED if v["c134"] > 0 else _WHITE, bold_val=True)
    r += 1
    _single(ws, r, 137,
            "Total saldo a favor  (130 + 131 + 132 − 129 − 133 − 135)", v["c137"],
            bg=_BG_GREEN if v["c137"] > 0 else _WHITE, bold_val=True)
    r += 1
    _blank(ws, r)
    r += 1

    # ── DATOS ADICIONALES ─────────────────────────────────────────────────────
    _hdr_full(ws, r, "DATOS ADICIONALES", _MED, height=15)
    r += 1
    dep_count = int(declaracion.get("dependientes") or 0)
    tipo_gan = declaracion.get("tipo_ganancia") or "—"
    _pair(ws, r,
          138, "Número de dependientes económicos", dep_count,
          139, "Adición por dependientes a casilla 92", 0)
    r += 1
    _pair(ws, r,
          140, "Superó tope indicativo Art. 336-1 ET (X=sí)", "—",
          141, "Tipo de ganancia ocasional", tipo_gan)
    r += 1
    _blank(ws, r)
    r += 1

    # Pie de página
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(row=r, column=1,
                value="Generado por TaxOps  —  Sólo para referencia interna  —  "
                      "No constituye declaración formal ante la DIAN")
    c.font = Font(name="Calibri", size=8, italic=True, color="808080")
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[r].height = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
