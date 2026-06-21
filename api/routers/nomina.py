"""Router de nómina — cálculo mensual, liquidación definitiva, empleados DB y períodos batch."""
from __future__ import annotations

import io
import json
import os
from datetime import date
from uuid import uuid4

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse

from dependencies import get_current_user
from services.nomina import (
    ARL_LABELS,
    ARL_RATES,
    EntradaLiquidacion,
    EntradaNomina,
    ResultadoLiquidacion,
    ResultadoNomina,
    SMLMV,
    AUX_TRANSPORTE,
    TOPE_AUX,
    SALARIO_INTEGRAL,
    calcular_liquidacion,
    calcular_nomina,
)

router = APIRouter(prefix="/nomina", tags=["Nómina"])


@router.get("/parametros")
async def parametros(_: dict = Depends(get_current_user)) -> dict:
    """Retorna parámetros legales 2026 y tasas ARL para el frontend."""
    return {
        "smlmv": SMLMV,
        "aux_transporte": AUX_TRANSPORTE,
        "tope_aux_transporte": TOPE_AUX,
        "salario_integral_minimo": SALARIO_INTEGRAL,
        "arl_clases": ARL_LABELS,
        "arl_rates": ARL_RATES,
        "ano": 2026,
    }


# ── Helpers internos ──────────────────────────────────────────────────────────

def _mensual_body_to_entrada(body: dict) -> EntradaNomina:
    return EntradaNomina(
        salario_basico              = float(body.get("salario_basico", 0)),
        dias_trabajados             = int(body.get("dias_trabajados", 30)),
        hed                         = float(body.get("hed", 0)),
        hen                         = float(body.get("hen", 0)),
        rn                          = float(body.get("rn", 0)),
        hod                         = float(body.get("hod", 0)),
        hedd                        = float(body.get("hedd", 0)),
        hend                        = float(body.get("hend", 0)),
        rnd                         = float(body.get("rnd", 0)),
        bonificaciones_salariales   = float(body.get("bonificaciones_salariales", 0)),
        comisiones                  = float(body.get("comisiones", 0)),
        otros_no_salariales         = float(body.get("otros_no_salariales", 0)),
        retencion_fuente            = float(body.get("retencion_fuente", 0)),
        otras_deducciones           = float(body.get("otras_deducciones", 0)),
        clase_riesgo_arl            = int(body.get("clase_riesgo_arl", 1)),
        tipo_salario                = body.get("tipo_salario", "ordinario"),
    )


def _mensual_to_dict(r: ResultadoNomina) -> dict:
    return {
        "devengado": {
            "salario_proporcional":   r.salario_proporcional,
            "auxilio_transporte":     r.auxilio_transporte,
            "aplica_aux_transporte":  r.aplica_aux_transporte,
            "horas_extras": {
                "hed":  r.he_hed,
                "hen":  r.he_hen,
                "rn":   r.he_rn,
                "hod":  r.he_hod,
                "hedd": r.he_hedd,
                "hend": r.he_hend,
                "rnd":  r.he_rnd,
                "subtotal": r.subtotal_horas_extras,
            },
            "bonificaciones":         r.bonificaciones,
            "comisiones":             r.comisiones,
            "otros_no_salariales":    r.otros_no_salariales,
            "total_devengado":        r.total_devengado,
        },
        "deducciones": {
            "ibc":               r.ibc,
            "salud_empleado":    r.salud_empleado,
            "pension_empleado":  r.pension_empleado,
            "fsp":               r.fsp,
            "retencion_fuente":  r.retencion_fuente,
            "otras":             r.otras_deducciones,
            "total_deducciones": r.total_deducciones,
        },
        "carga_patronal": {
            "pension":             r.pension_empleador,
            "salud_empleador":     r.salud_empleador,
            "arl":                 r.arl,
            "sena":                r.sena,
            "icbf":                r.icbf,
            "caja_compensacion":   r.caja_compensacion,
            "exonerado_sena_icbf": r.exonerado_sena_icbf,
            "total":               r.total_carga_patronal,
            "costo_total_empresa": r.costo_total_empresa,
        },
        "neto_pagar": r.neto_pagar,
    }


@router.post("/mensual")
async def calcular_mensual(
    body: dict,
    _: dict = Depends(get_current_user),
) -> dict:
    entrada = _mensual_body_to_entrada(body)
    return _mensual_to_dict(calcular_nomina(entrada))


@router.post("/liquidacion")
async def calcular_liq(
    body: dict,
    _: dict = Depends(get_current_user),
) -> dict:
    """Calcula liquidación definitiva de contrato."""
    def _d(key: str) -> date:
        v = body.get(key, "")
        if isinstance(v, date):
            return v
        return date.fromisoformat(str(v)[:10])

    entrada = EntradaLiquidacion(
        salario_basico                = float(body.get("salario_basico", 0)),
        fecha_ingreso                 = _d("fecha_ingreso"),
        fecha_retiro                  = _d("fecha_retiro"),
        tipo_contrato                 = body.get("tipo_contrato", "indefinido"),
        causa_terminacion             = body.get("causa_terminacion", "renuncia"),
        promedio_he_recargos          = float(body.get("promedio_he_recargos", 0)),
        promedio_bonif_salariales     = float(body.get("promedio_bonif_salariales", 0)),
        dias_vacaciones_disfrutadas   = float(body.get("dias_vacaciones_disfrutadas", 0)),
        anticipo_cesantias            = float(body.get("anticipo_cesantias", 0)),
        otras_deducciones             = float(body.get("otras_deducciones", 0)),
        tipo_salario                  = body.get("tipo_salario", "ordinario"),
    )
    r = calcular_liquidacion(entrada)
    return {
        "tiempo": {
            "fecha_ingreso":   body.get("fecha_ingreso"),
            "fecha_retiro":    body.get("fecha_retiro"),
            "dias_trabajados": r.dias_trabajados,
        },
        "bases": {
            "salario_base_prestacional":  r.salario_base_prestacional,
            "salario_base_vacaciones":    r.salario_base_vacaciones,
            "aplica_aux_transporte":      r.aplica_aux_transporte,
        },
        "prestaciones": {
            "cesantias":              r.cesantias,
            "intereses_cesantias":    r.intereses_cesantias,
            "prima_servicios":        r.prima_servicios,
            "vacaciones":             r.vacaciones_proporcionales,
            "salario_pendiente":      r.salario_pendiente,
        },
        "indemnizacion": {
            "valor":  r.indemnizacion,
            "detalle": r.causa_indemnizacion,
        },
        "dias_auditoria": {
            "cesantias":               r.dias_cesantias,
            "prima":                   r.dias_prima,
            "vacaciones_causadas":     r.dias_vacaciones_causadas,
            "vacaciones_a_pagar":      r.dias_vacaciones_a_pagar,
        },
        "totales": {
            "subtotal_devengado": r.subtotal_devengado,
            "deducciones":        r.otras_deducciones,
            "total_neto":         r.total_neto,
        },
    }


# ── Exportación Excel ─────────────────────────────────────────────────────────

@router.post("/export")
async def exportar(
    body: dict,
    _: dict = Depends(get_current_user),
) -> StreamingResponse:
    """
    Genera un archivo .xlsx con el resultado de nómina mensual o liquidación.
    Body: { tipo: 'mensual'|'liquidacion', ...mismos campos del endpoint correspondiente }
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    tipo = body.get("tipo", "mensual")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nómina"

    NAVY = "001B3A"; ORANGE = "E85D04"; LIGHT = "EEF2F7"; WHITE = "FFFFFF"

    def section(row: int, title: str) -> None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row=row, column=1, value=title)
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=ORANGE)

    def row_data(r: int, label: str, value, bold: bool = False) -> None:
        lc = ws.cell(row=r, column=1, value=label)
        vc = ws.cell(row=r, column=3, value=value)
        if bold:
            for col in [1, 3]:
                ws.cell(row=r, column=col).font = Font(bold=True)
            for col in [1, 2, 3]:
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=LIGHT)
        if isinstance(value, (int, float)):
            vc.number_format = '"$"#,##0'
            vc.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 3
    ws.column_dimensions["C"].width = 18

    # Título
    ws.merge_cells("A1:C1")
    t = ws.cell(row=1, column=1, value="TaxOps · Liquidación de Nómina CST Colombia 2026")
    t.font = Font(bold=True, size=13, color=WHITE)
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:C2")
    sub = ws.cell(row=2, column=1,
        value=f"{'Nómina Mensual' if tipo == 'mensual' else 'Liquidación Definitiva'} · SMLMV $1,750,905 · Aux.Transporte $249,095")
    sub.font = Font(italic=True, size=9, color="444444")
    sub.alignment = Alignment(horizontal="center")

    row = 4

    if tipo == "mensual":
        r = calcular_nomina(_mensual_body_to_entrada(body))

        section(row, "▸ DEVENGADO"); row += 1
        row_data(row, "Salario proporcional", r.salario_proporcional); row += 1
        if r.aplica_aux_transporte:
            row_data(row, "Auxilio de transporte", r.auxilio_transporte); row += 1
        he_labels = [
            ("Extra Diurna HED (+25%)", r.he_hed),
            ("Extra Nocturna HEN (+75%)", r.he_hen),
            ("Recargo Nocturno RN (+35%)", r.he_rn),
            ("Dominical/Festivo HOD (+75%)", r.he_hod),
            ("Extra Diurna Dom. HEDD (+100%)", r.he_hedd),
            ("Extra Nocturna Dom. HEND (+150%)", r.he_hend),
            ("Recargo Noct. Dom. RND (+110%)", r.he_rnd),
        ]
        for label, val in he_labels:
            if val:
                row_data(row, f"  · {label}", val); row += 1
        if r.bonificaciones:
            row_data(row, "Bonificaciones salariales", r.bonificaciones); row += 1
        if r.comisiones:
            row_data(row, "Comisiones", r.comisiones); row += 1
        if r.otros_no_salariales:
            row_data(row, "Conceptos no salariales", r.otros_no_salariales); row += 1
        row_data(row, "TOTAL DEVENGADO", r.total_devengado, bold=True); row += 2

        section(row, "▸ DEDUCCIONES EMPLEADO"); row += 1
        row_data(row, f"IBC – Base Seg. Social", r.ibc); row += 1
        row_data(row, "Salud empleado (4%)", r.salud_empleado); row += 1
        row_data(row, "Pensión empleado (4%)", r.pension_empleado); row += 1
        if r.fsp:
            row_data(row, "FSP – Fondo Solidaridad Pensional", r.fsp); row += 1
        if r.retencion_fuente:
            row_data(row, "Retención en la fuente", r.retencion_fuente); row += 1
        if r.otras_deducciones:
            row_data(row, "Otras deducciones", r.otras_deducciones); row += 1
        row_data(row, "TOTAL DEDUCCIONES", r.total_deducciones, bold=True); row += 2

        ws.merge_cells(f"A{row}:C{row}")
        nc = ws.cell(row=row, column=1, value=f"NETO A PAGAR: ${r.neto_pagar:,}")
        nc.font = Font(bold=True, size=13, color=WHITE)
        nc.fill = PatternFill("solid", fgColor=NAVY)
        nc.alignment = Alignment(horizontal="center")
        ws.row_dimensions[row].height = 24; row += 2

        exon_txt = " (EXONERADO Art.114-1 ET)" if r.exonerado_sena_icbf else ""
        arl_clase = int(body.get("clase_riesgo_arl", 1))
        section(row, "▸ CARGA PATRONAL – PARAFISCALES Y SEGURIDAD SOCIAL"); row += 1
        row_data(row, "Salud empleador (8.5%)", r.salud_empleador); row += 1
        row_data(row, "Pensión empleador (12%)", r.pension_empleador); row += 1
        row_data(row, f"ARL – {ARL_LABELS.get(arl_clase, ARL_LABELS[1])}", r.arl); row += 1
        row_data(row, f"SENA (2%){exon_txt}", r.sena); row += 1
        row_data(row, f"ICBF (3%){exon_txt}", r.icbf); row += 1
        row_data(row, "Caja de Compensación Familiar (4%)", r.caja_compensacion); row += 1
        row_data(row, "TOTAL CARGA PATRONAL", r.total_carga_patronal, bold=True); row += 2

        ws.merge_cells(f"A{row}:C{row}")
        ec = ws.cell(row=row, column=1, value=f"COSTO TOTAL EMPRESA: ${r.costo_total_empresa:,}")
        ec.font = Font(bold=True, size=12, color=WHITE)
        ec.fill = PatternFill("solid", fgColor=ORANGE)
        ec.alignment = Alignment(horizontal="center")
        ws.row_dimensions[row].height = 22; row += 2

    else:
        def _d(key: str) -> date:
            v = body.get(key, "")
            return date.fromisoformat(str(v)[:10]) if not isinstance(v, date) else v

        entrada = EntradaLiquidacion(
            salario_basico              = float(body.get("salario_basico", 0)),
            fecha_ingreso               = _d("fecha_ingreso"),
            fecha_retiro                = _d("fecha_retiro"),
            tipo_contrato               = body.get("tipo_contrato", "indefinido"),
            causa_terminacion           = body.get("causa_terminacion", "renuncia"),
            promedio_he_recargos        = float(body.get("promedio_he_recargos", 0)),
            promedio_bonif_salariales   = float(body.get("promedio_bonif_salariales", 0)),
            dias_vacaciones_disfrutadas = float(body.get("dias_vacaciones_disfrutadas", 0)),
            anticipo_cesantias          = float(body.get("anticipo_cesantias", 0)),
            otras_deducciones           = float(body.get("otras_deducciones", 0)),
            tipo_salario                = body.get("tipo_salario", "ordinario"),
        )
        r2 = calcular_liquidacion(entrada)

        section(row, "▸ DATOS DEL CONTRATO"); row += 1
        row_data(row, "Fecha de ingreso", str(body.get("fecha_ingreso", ""))); row += 1
        row_data(row, "Fecha de retiro", str(body.get("fecha_retiro", ""))); row += 1
        row_data(row, "Días trabajados (base 360)", r2.dias_trabajados); row += 1
        row_data(row, "Tipo de contrato", body.get("tipo_contrato", "")); row += 1
        row_data(row, "Causa terminación", body.get("causa_terminacion", "")); row += 2

        section(row, "▸ PRESTACIONES SOCIALES"); row += 1
        row_data(row, f"Cesantías ({r2.dias_cesantias} días)", r2.cesantias); row += 1
        row_data(row, "Intereses sobre cesantías (12%/360)", r2.intereses_cesantias); row += 1
        row_data(row, f"Prima de servicios ({r2.dias_prima} días)", r2.prima_servicios); row += 1
        row_data(row, f"Vacaciones ({round(r2.dias_vacaciones_a_pagar, 1)} días pendientes)", r2.vacaciones_proporcionales); row += 1
        row_data(row, "Salario pendiente último período", r2.salario_pendiente); row += 1
        if r2.indemnizacion:
            row_data(row, f"Indemnización · {r2.causa_indemnizacion}", r2.indemnizacion); row += 1
        row_data(row, "SUBTOTAL DEVENGADO", r2.subtotal_devengado, bold=True); row += 1
        if r2.otras_deducciones:
            row_data(row, "(-) Deducciones / préstamos", -r2.otras_deducciones); row += 1
        row += 1

        ws.merge_cells(f"A{row}:C{row}")
        nc = ws.cell(row=row, column=1, value=f"TOTAL NETO A PAGAR: ${r2.total_neto:,}")
        nc.font = Font(bold=True, size=13, color=WHITE)
        nc.fill = PatternFill("solid", fgColor=NAVY)
        nc.alignment = Alignment(horizontal="center")
        ws.row_dimensions[row].height = 24; row += 2

    row_data(row, "Generado por TaxOps SaaS", f"Fecha: {date.today().isoformat()}"); row += 1
    row_data(row, "Marco legal", "CST Colombia · Art.114-1 ET · Ley 1607/2012 · Ley 2101/2021")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"taxops_nomina_{tipo}_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Exportación PDF ───────────────────────────────────────────────────────────

@router.post("/export-pdf")
async def exportar_pdf(
    body: dict,
    _: dict = Depends(get_current_user),
) -> StreamingResponse:
    """Genera un PDF compacto del colado de nómina mensual o liquidación definitiva."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    tipo = body.get("tipo", "mensual")
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=1.8*cm, leftMargin=1.8*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    NAV = colors.HexColor("#1A3A5C")
    ORG = colors.HexColor("#E05519")
    LGT = colors.HexColor("#EEF2F7")
    styles = getSampleStyleSheet()
    H1 = ParagraphStyle("h1", parent=styles["Normal"], fontSize=15, textColor=colors.white,
                        backColor=NAV, spaceAfter=2, alignment=TA_CENTER, fontName="Helvetica-Bold", leading=20)
    H2 = ParagraphStyle("h2", parent=styles["Normal"], fontSize=9, textColor=colors.white,
                        backColor=ORG, spaceAfter=1, spaceBefore=8, fontName="Helvetica-Bold", leading=13)
    FOOT = ParagraphStyle("foot", parent=styles["Normal"], fontSize=7, textColor=colors.grey, alignment=TA_CENTER)

    def make_table(rows):
        t = Table(rows, colWidths=[10*cm, 4.5*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), NAV),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LGT]),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#DDDDDD")),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        return t

    def cop(v): return f"${v:,.0f}"

    story = []
    story.append(Paragraph("TaxOps · Nómina CST Colombia 2026", H1))
    subtitle = "Nómina Mensual" if tipo == "mensual" else "Liquidación Definitiva de Contrato"
    story.append(Paragraph(f"{subtitle}  ·  SMLMV $1,750,905  ·  Aux. Transporte $249,095", styles["Normal"]))
    story.append(Spacer(1, 0.3*cm))

    if tipo == "mensual":
        r = calcular_nomina(_mensual_body_to_entrada(body))

        story.append(Paragraph("DEVENGADO", H2))
        dev_rows = [["Concepto", "Valor"]]
        dev_rows.append(["Salario proporcional", cop(r.salario_proporcional)])
        if r.aplica_aux_transporte:
            dev_rows.append(["Auxilio de transporte", cop(r.auxilio_transporte)])
        he_map = [("HED +25%", r.he_hed), ("HEN +75%", r.he_hen), ("RN +35%", r.he_rn),
                  ("HOD +75%", r.he_hod), ("HEDD +100%", r.he_hedd), ("HEND +150%", r.he_hend), ("RND +110%", r.he_rnd)]
        for lbl, v in he_map:
            if v: dev_rows.append([f"  · {lbl}", cop(v)])
        if r.bonificaciones: dev_rows.append(["Bonificaciones salariales", cop(r.bonificaciones)])
        if r.comisiones: dev_rows.append(["Comisiones", cop(r.comisiones)])
        if r.otros_no_salariales: dev_rows.append(["No salariales", cop(r.otros_no_salariales)])
        dev_rows.append(["TOTAL DEVENGADO", cop(r.total_devengado)])
        story.append(make_table(dev_rows))

        story.append(Paragraph("DEDUCCIONES EMPLEADO", H2))
        ded_rows = [["Concepto", "Valor"],
            [f"IBC – Base Seg. Social", cop(r.ibc)],
            ["Salud empleado (4%)", cop(r.salud_empleado)],
            ["Pensión empleado (4%)", cop(r.pension_empleado)]]
        if r.fsp: ded_rows.append(["FSP – Solidaridad Pensional", cop(r.fsp)])
        if r.retencion_fuente: ded_rows.append(["Retención en la fuente", cop(r.retencion_fuente)])
        if r.otras_deducciones: ded_rows.append(["Otras deducciones", cop(r.otras_deducciones)])
        ded_rows.append(["TOTAL DEDUCCIONES", cop(r.total_deducciones)])
        story.append(make_table(ded_rows))

        # Neto highlight
        story.append(Spacer(1, 0.2*cm))
        neto_t = Table([[f"NETO A PAGAR", cop(r.neto_pagar)]], colWidths=[10*cm, 4.5*cm])
        neto_t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), NAV),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 12),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(neto_t)

        story.append(Paragraph("CARGA PATRONAL – PARAFISCALES", H2))
        exon = " (EXONERADO)" if r.exonerado_sena_icbf else ""
        arl_clase = int(body.get("clase_riesgo_arl", 1))
        pat_rows = [["Concepto", "Valor"],
            ["Salud empleador (8.5%)", cop(r.salud_empleador)],
            ["Pensión empleador (12%)", cop(r.pension_empleador)],
            [f"ARL · {ARL_LABELS.get(arl_clase, '')}", cop(r.arl)],
            [f"SENA (2%){exon}", cop(r.sena)],
            [f"ICBF (3%){exon}", cop(r.icbf)],
            ["Caja de Compensación (4%)", cop(r.caja_compensacion)],
            ["TOTAL CARGA PATRONAL", cop(r.total_carga_patronal)]]
        story.append(make_table(pat_rows))

        story.append(Spacer(1, 0.2*cm))
        costo_t = Table([["COSTO TOTAL EMPRESA", cop(r.costo_total_empresa)]], colWidths=[10*cm, 4.5*cm])
        costo_t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), ORG),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 11),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(costo_t)

    else:
        def _d(key: str) -> date:
            v = body.get(key, "")
            return date.fromisoformat(str(v)[:10]) if not isinstance(v, date) else v

        entrada = EntradaLiquidacion(
            salario_basico              = float(body.get("salario_basico", 0)),
            fecha_ingreso               = _d("fecha_ingreso"),
            fecha_retiro                = _d("fecha_retiro"),
            tipo_contrato               = body.get("tipo_contrato", "indefinido"),
            causa_terminacion           = body.get("causa_terminacion", "renuncia"),
            promedio_he_recargos        = float(body.get("promedio_he_recargos", 0)),
            promedio_bonif_salariales   = float(body.get("promedio_bonif_salariales", 0)),
            dias_vacaciones_disfrutadas = float(body.get("dias_vacaciones_disfrutadas", 0)),
            anticipo_cesantias          = float(body.get("anticipo_cesantias", 0)),
            otras_deducciones           = float(body.get("otras_deducciones", 0)),
            tipo_salario                = body.get("tipo_salario", "ordinario"),
        )
        r2 = calcular_liquidacion(entrada)

        story.append(Paragraph("DATOS DEL CONTRATO", H2))
        info_rows = [["Campo", "Valor"],
            ["Fecha de ingreso", str(body.get("fecha_ingreso", ""))],
            ["Fecha de retiro", str(body.get("fecha_retiro", ""))],
            ["Días trabajados (base 360)", str(r2.dias_trabajados)],
            ["Tipo de contrato", body.get("tipo_contrato", "")],
            ["Causa terminación", body.get("causa_terminacion", "")]]
        story.append(make_table(info_rows))

        story.append(Paragraph("PRESTACIONES SOCIALES", H2))
        pres_rows = [["Concepto", "Valor"],
            [f"Cesantías ({r2.dias_cesantias} días)", cop(r2.cesantias)],
            ["Intereses sobre cesantías (12%)", cop(r2.intereses_cesantias)],
            [f"Prima de servicios ({r2.dias_prima} días)", cop(r2.prima_servicios)],
            [f"Vacaciones ({round(r2.dias_vacaciones_a_pagar,1)} días pendientes)", cop(r2.vacaciones_proporcionales)],
            ["Salario pendiente", cop(r2.salario_pendiente)]]
        if r2.indemnizacion:
            pres_rows.append([f"Indemnización · {r2.causa_indemnizacion}", cop(r2.indemnizacion)])
        pres_rows.append(["SUBTOTAL DEVENGADO", cop(r2.subtotal_devengado)])
        if r2.otras_deducciones:
            pres_rows.append(["(-) Deducciones / préstamos", cop(-r2.otras_deducciones)])
        story.append(make_table(pres_rows))

        story.append(Spacer(1, 0.2*cm))
        neto_t = Table([["TOTAL NETO A PAGAR", cop(r2.total_neto)]], colWidths=[10*cm, 4.5*cm])
        neto_t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), NAV),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 12),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(neto_t)

    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(
        f"Generado por TaxOps SaaS · {date.today().strftime('%d/%m/%Y')} · "
        "Marco legal: CST Colombia · Art.114-1 ET · Ley 1607/2012 · Ley 2101/2021",
        FOOT,
    ))

    doc.build(story)
    buf.seek(0)
    filename = f"taxops_nomina_{tipo}_{date.today().isoformat()}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})


# ── Import batch desde Excel/CSV ──────────────────────────────────────────────

@router.post("/import-batch")
async def import_batch(
    file: UploadFile = File(...),
    _: dict = Depends(get_current_user),
) -> list:
    """
    Recibe un Excel (.xlsx/.xls) o CSV con columnas:
      nombre, salario_basico, [dias_trabajados], [clase_riesgo_arl], [tipo_salario]
    Calcula nómina mensual para cada fila y devuelve resultados.
    """
    import pandas as pd

    content = await file.read()
    filename = (file.filename or "").lower()

    if filename.endswith(".csv"):
        df = pd.read_csv(io.BytesIO(content))
    else:
        df = pd.read_excel(io.BytesIO(content))

    # Normalizar columnas
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]

    results = []
    for _, row in df.iterrows():
        try:
            salario = float(str(row.get("salario_basico", 0)).replace(".", "").replace(",", ".").replace("$", "").strip() or 0)
            entrada = EntradaNomina(
                salario_basico    = salario,
                dias_trabajados   = int(row.get("dias_trabajados", 30)),
                clase_riesgo_arl  = int(row.get("clase_riesgo_arl", 1)),
                tipo_salario      = str(row.get("tipo_salario", "ordinario")),
            )
            r = calcular_nomina(entrada)
            results.append({
                "nombre":  str(row.get("nombre", f"Fila {_ + 1}")),
                "salario": salario,
                "neto":    r.neto_pagar,
                "carga":   r.total_carga_patronal,
                "costo":   r.costo_total_empresa,
            })
        except Exception as exc:
            results.append({
                "nombre": str(row.get("nombre", "?")),
                "salario": 0, "neto": 0, "carga": 0, "costo": 0,
                "error": str(exc),
            })

    return results


# ── Export batch ──────────────────────────────────────────────────────────────

@router.post("/export-batch")
async def export_batch(
    body: dict,
    _: dict = Depends(get_current_user),
) -> StreamingResponse:
    """Genera Excel con los resultados del batch import."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = body.get("rows", [])
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nómina Batch"

    NAVY = "001B3A"; ORANGE = "E85D04"; WHITE = "FFFFFF"; LIGHT = "EEF2F7"
    headers = ["Nombre", "Salario Básico", "Neto a Pagar", "Carga Patronal", "Costo Total Empresa"]
    widths = [30, 18, 18, 18, 22]

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[chr(64 + i)].width = w

    for ri, r in enumerate(rows, 2):
        ws.cell(ri, 1, r.get("nombre", "")).font = Font(size=9)
        for ci, key in enumerate(["salario","neto","carga","costo"], 2):
            c = ws.cell(ri, ci, r.get(key, 0))
            c.number_format = '"$"#,##0'
            c.alignment = Alignment(horizontal="right")
            if ri % 2 == 0:
                c.fill = PatternFill("solid", fgColor=LIGHT)

    # Totals row
    tr = len(rows) + 2
    ws.cell(tr, 1, "TOTALES").font = Font(bold=True)
    for ci, key in enumerate(["salario","neto","carga","costo"], 2):
        total = sum(r.get(key, 0) for r in rows)
        c = ws.cell(tr, ci, total)
        c.font = Font(bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=ORANGE)
        c.number_format = '"$"#,##0'
        c.alignment = Alignment(horizontal="right")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"taxops_nomina_batch_{date.today().isoformat()}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'})


# ═══════════════════════════════════════════════════════════════════════════════
# EMPLEADOS — CRUD persistido en PostgreSQL
# ═══════════════════════════════════════════════════════════════════════════════

def _get_db():
    from db.database import db_available, get_db
    if not db_available():
        raise HTTPException(status_code=503, detail="Base de datos no disponible")
    return get_db


@router.get("/empleados")
async def listar_empleados(user: dict = Depends(get_current_user)) -> list:
    from sqlalchemy import text
    get_db = _get_db()
    with get_db() as db:
        rows = db.execute(
            text("SELECT * FROM empleados WHERE org_id = :org AND activo = TRUE ORDER BY nombre"),
            {"org": user["org_id"]},
        ).mappings().fetchall()
    return [dict(r) for r in rows]


@router.post("/empleados", status_code=201)
async def crear_empleado(body: dict, user: dict = Depends(get_current_user)) -> dict:
    from sqlalchemy import text
    if not body.get("nombre") or not body.get("salario"):
        raise HTTPException(status_code=422, detail="nombre y salario son requeridos")
    get_db = _get_db()
    emp_id = str(uuid4())
    with get_db() as db:
        db.execute(
            text("""
                INSERT INTO empleados
                    (id, org_id, cedula, nombre, cargo, salario,
                     clase_riesgo_arl, tipo_salario, fecha_ingreso, email)
                VALUES
                    (:id, :org, :cedula, :nombre, :cargo, :salario,
                     :arl, :tipo, :fecha, :email)
            """),
            {
                "id":     emp_id,
                "org":    user["org_id"],
                "cedula": body.get("cedula", ""),
                "nombre": body["nombre"],
                "cargo":  body.get("cargo", ""),
                "salario": float(body["salario"]),
                "arl":    int(body.get("clase_riesgo_arl", 1)),
                "tipo":   body.get("tipo_salario", "ordinario"),
                "fecha":  body.get("fecha_ingreso") or None,
                "email":  body.get("email", ""),
            },
        )
        db.commit()
    return {"id": emp_id, **body}


@router.put("/empleados/{emp_id}")
async def actualizar_empleado(emp_id: str, body: dict, user: dict = Depends(get_current_user)) -> dict:
    from sqlalchemy import text
    get_db = _get_db()
    with get_db() as db:
        result = db.execute(
            text("""
                UPDATE empleados SET
                    cedula = :cedula, nombre = :nombre, cargo = :cargo,
                    salario = :salario, clase_riesgo_arl = :arl,
                    tipo_salario = :tipo, fecha_ingreso = :fecha,
                    email = :email, updated_at = NOW()
                WHERE id = :id AND org_id = :org AND activo = TRUE
            """),
            {
                "id":     emp_id,
                "org":    user["org_id"],
                "cedula": body.get("cedula", ""),
                "nombre": body.get("nombre", ""),
                "cargo":  body.get("cargo", ""),
                "salario": float(body.get("salario", 0)),
                "arl":    int(body.get("clase_riesgo_arl", 1)),
                "tipo":   body.get("tipo_salario", "ordinario"),
                "fecha":  body.get("fecha_ingreso") or None,
                "email":  body.get("email", ""),
            },
        )
        db.commit()
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="Empleado no encontrado")
    return {"id": emp_id, **body}


@router.delete("/empleados/{emp_id}", status_code=204)
async def eliminar_empleado(emp_id: str, user: dict = Depends(get_current_user)) -> None:
    from sqlalchemy import text
    get_db = _get_db()
    with get_db() as db:
        result = db.execute(
            text("UPDATE empleados SET activo = FALSE WHERE id = :id AND org_id = :org"),
            {"id": emp_id, "org": user["org_id"]},
        )
        db.commit()
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="Empleado no encontrado")


# ═══════════════════════════════════════════════════════════════════════════════
# PERÍODOS — cálculo batch + historial en DB
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/periodos")
async def listar_periodos(user: dict = Depends(get_current_user)) -> list:
    from sqlalchemy import text
    get_db = _get_db()
    with get_db() as db:
        rows = db.execute(
            text("SELECT * FROM nomina_periodos WHERE org_id = :org ORDER BY periodo DESC"),
            {"org": user["org_id"]},
        ).mappings().fetchall()
    return [dict(r) for r in rows]


@router.get("/periodos/{periodo}")
async def detalle_periodo(periodo: str, user: dict = Depends(get_current_user)) -> dict:
    from sqlalchemy import text
    get_db = _get_db()
    with get_db() as db:
        per = db.execute(
            text("SELECT * FROM nomina_periodos WHERE org_id = :org AND periodo = :p LIMIT 1"),
            {"org": user["org_id"], "p": periodo},
        ).mappings().fetchone()
        if not per:
            raise HTTPException(status_code=404, detail="Período no encontrado")
        detalle = db.execute(
            text("""
                SELECT d.*, e.nombre, e.cedula, e.cargo, e.email
                FROM nomina_detalle d
                JOIN empleados e ON e.id = d.empleado_id
                WHERE d.periodo_id = :pid
                ORDER BY e.nombre
            """),
            {"pid": per["id"]},
        ).mappings().fetchall()
    return {
        **dict(per),
        "detalle": [dict(r) for r in detalle],
    }


@router.post("/periodos/{periodo}/calcular")
async def calcular_periodo(periodo: str, body: dict, user: dict = Depends(get_current_user)) -> dict:
    """Calcula nómina batch para todos los empleados activos del período YYYY-MM."""
    import re
    if not re.match(r"^\d{4}-\d{2}$", periodo):
        raise HTTPException(status_code=422, detail="Formato de período inválido. Use YYYY-MM")

    from sqlalchemy import text
    get_db = _get_db()
    org_id = user["org_id"]

    with get_db() as db:
        empleados_rows = db.execute(
            text("SELECT * FROM empleados WHERE org_id = :org AND activo = TRUE ORDER BY nombre"),
            {"org": org_id},
        ).mappings().fetchall()

    if not empleados_rows:
        raise HTTPException(status_code=422, detail="No hay empleados activos en esta organización")

    dias = int(body.get("dias_trabajados", 30))
    resultados = []
    total_dev = total_ded = total_neto = total_carga = total_costo = 0.0

    for emp in empleados_rows:
        try:
            entrada = EntradaNomina(
                salario_basico   = float(emp["salario"]),
                dias_trabajados  = dias,
                clase_riesgo_arl = int(emp["clase_riesgo_arl"]),
                tipo_salario     = emp["tipo_salario"],
            )
            r = calcular_nomina(entrada)
            row = {
                "empleado_id":   str(emp["id"]),
                "nombre":        emp["nombre"],
                "cedula":        emp["cedula"],
                "cargo":         emp["cargo"],
                "email":         emp["email"],
                "salario":       float(emp["salario"]),
                "devengado":     r.total_devengado,
                "deducciones":   r.total_deducciones,
                "neto":          r.neto_pagar,
                "carga":         r.total_carga_patronal,
                "costo":         r.costo_total_empresa,
                "detalle":       _mensual_to_dict(r),
            }
            total_dev   += r.total_devengado
            total_ded   += r.total_deducciones
            total_neto  += r.neto_pagar
            total_carga += r.total_carga_patronal
            total_costo += r.costo_total_empresa
        except Exception as exc:
            row = {
                "empleado_id": str(emp["id"]),
                "nombre": emp["nombre"],
                "cedula": emp["cedula"],
                "error":  str(exc),
                "devengado": 0, "deducciones": 0, "neto": 0, "carga": 0, "costo": 0,
            }
        resultados.append(row)

    # Persistir o actualizar el período
    with get_db() as db:
        existing = db.execute(
            text("SELECT id FROM nomina_periodos WHERE org_id = :org AND periodo = :p AND tipo = 'mensual'"),
            {"org": org_id, "p": periodo},
        ).fetchone()

        if existing:
            periodo_id = str(existing[0])
            db.execute(
                text("""
                    UPDATE nomina_periodos SET
                        total_devengado = :dev, total_deducciones = :ded,
                        total_neto = :neto, carga_patronal = :carga,
                        costo_empresa = :costo, empleados_count = :cnt,
                        estado = 'calculado'
                    WHERE id = :pid
                """),
                {"pid": periodo_id, "dev": total_dev, "ded": total_ded,
                 "neto": total_neto, "carga": total_carga, "costo": total_costo,
                 "cnt": len(resultados)},
            )
            db.execute(text("DELETE FROM nomina_detalle WHERE periodo_id = :pid"), {"pid": periodo_id})
        else:
            periodo_id = str(uuid4())
            db.execute(
                text("""
                    INSERT INTO nomina_periodos
                        (id, org_id, periodo, tipo, estado,
                         total_devengado, total_deducciones, total_neto,
                         carga_patronal, costo_empresa, empleados_count)
                    VALUES
                        (:id, :org, :p, 'mensual', 'calculado',
                         :dev, :ded, :neto, :carga, :costo, :cnt)
                """),
                {"id": periodo_id, "org": org_id, "p": periodo,
                 "dev": total_dev, "ded": total_ded, "neto": total_neto,
                 "carga": total_carga, "costo": total_costo, "cnt": len(resultados)},
            )

        for row in resultados:
            if "error" in row:
                continue
            db.execute(
                text("""
                    INSERT INTO nomina_detalle (id, periodo_id, empleado_id, org_id, resultado)
                    VALUES (:id, :pid, :eid, :org, :res)
                """),
                {
                    "id":  str(uuid4()),
                    "pid": periodo_id,
                    "eid": row["empleado_id"],
                    "org": org_id,
                    "res": json.dumps(row["detalle"], default=str),
                },
            )
        db.commit()

    return {
        "periodo_id":       periodo_id,
        "periodo":          periodo,
        "empleados_count":  len(resultados),
        "total_devengado":  total_dev,
        "total_deducciones": total_ded,
        "total_neto":       total_neto,
        "carga_patronal":   total_carga,
        "costo_empresa":    total_costo,
        "detalle":          resultados,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# COLILLA PDF individual por empleado/período
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/colilla/{periodo}/{empleado_id}")
async def colilla_pdf(periodo: str, empleado_id: str, user: dict = Depends(get_current_user)) -> StreamingResponse:
    """Genera colilla PDF para un empleado en un período calculado."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from sqlalchemy import text

    get_db = _get_db()
    with get_db() as db:
        row = db.execute(
            text("""
                SELECT d.resultado, e.nombre, e.cedula, e.cargo, e.salario, e.email
                FROM nomina_detalle d
                JOIN empleados e ON e.id = d.empleado_id
                JOIN nomina_periodos p ON p.id = d.periodo_id
                WHERE p.org_id = :org AND p.periodo = :per AND d.empleado_id = :eid
                LIMIT 1
            """),
            {"org": user["org_id"], "per": periodo, "eid": empleado_id},
        ).mappings().fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Colilla no encontrada. Calcula el período primero.")

    resultado = row["resultado"] if isinstance(row["resultado"], dict) else json.loads(row["resultado"])
    emp_nombre = row["nombre"]
    emp_cedula = row["cedula"]
    emp_cargo  = row["cargo"]

    NAV = colors.HexColor("#1A3A5C")
    ORG = colors.HexColor("#E05519")
    LGT = colors.HexColor("#EEF2F7")
    styles = getSampleStyleSheet()
    H1 = ParagraphStyle("h1", parent=styles["Normal"], fontSize=14, textColor=colors.white,
                        backColor=NAV, alignment=TA_CENTER, fontName="Helvetica-Bold", leading=20, spaceAfter=2)
    H2 = ParagraphStyle("h2", parent=styles["Normal"], fontSize=9, textColor=colors.white,
                        backColor=ORG, spaceAfter=1, spaceBefore=6, fontName="Helvetica-Bold", leading=13)
    FOOT = ParagraphStyle("foot", parent=styles["Normal"], fontSize=7, textColor=colors.grey, alignment=TA_CENTER)

    def make_table(rows):
        t = Table(rows, colWidths=[10*cm, 4.5*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), NAV),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LGT]),
            ("GRID",       (0, 0), (-1, -1), 0.25, colors.HexColor("#DDDDDD")),
            ("ALIGN",      (1, 0), (1, -1), "RIGHT"),
            ("LEFTPADDING",  (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING",   (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
        ]))
        return t

    def cop(v): return f"${float(v or 0):,.0f}"

    dev  = resultado.get("devengado", {})
    ded  = resultado.get("deducciones", {})
    pat  = resultado.get("carga_patronal", {})
    neto = resultado.get("neto_pagar", 0)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=1.8*cm, leftMargin=1.8*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    story = []

    story.append(Paragraph(f"TaxOps · Colilla de Nómina · {periodo}", H1))
    story.append(Paragraph(f"Empleado: {emp_nombre}  ·  CC {emp_cedula}  ·  {emp_cargo}", styles["Normal"]))
    story.append(Spacer(1, 0.3*cm))

    story.append(Paragraph("DEVENGADO", H2))
    dev_rows = [["Concepto", "Valor"],
        ["Salario proporcional", cop(dev.get("salario_proporcional", 0))]]
    if dev.get("aplica_aux_transporte"):
        dev_rows.append(["Auxilio de transporte", cop(dev.get("auxilio_transporte", 0))])
    he = dev.get("horas_extras", {})
    for lbl, key in [("HED +25%","hed"),("HEN +80%","hen"),("RN +35%","rn"),
                     ("HOD +80%","hod"),("HEDD +105%","hedd"),("HEND +155%","hend"),("RND +115%","rnd")]:
        if he.get(key, 0):
            dev_rows.append([f"  · {lbl}", cop(he[key])])
    if dev.get("bonificaciones"): dev_rows.append(["Bonificaciones salariales", cop(dev["bonificaciones"])])
    if dev.get("comisiones"):     dev_rows.append(["Comisiones", cop(dev["comisiones"])])
    if dev.get("otros_no_salariales"): dev_rows.append(["No salariales", cop(dev["otros_no_salariales"])])
    dev_rows.append(["TOTAL DEVENGADO", cop(dev.get("total_devengado", 0))])
    story.append(make_table(dev_rows))

    story.append(Paragraph("DEDUCCIONES EMPLEADO", H2))
    ded_rows = [["Concepto", "Valor"],
        ["IBC – Base Seg. Social",     cop(ded.get("ibc", 0))],
        ["Salud empleado (4%)",         cop(ded.get("salud_empleado", 0))],
        ["Pensión empleado (4%)",        cop(ded.get("pension_empleado", 0))]]
    if ded.get("fsp"):               ded_rows.append(["FSP – Solidaridad",    cop(ded["fsp"])])
    if ded.get("retencion_fuente"):  ded_rows.append(["Retención en la fuente", cop(ded["retencion_fuente"])])
    if ded.get("otras"):             ded_rows.append(["Otras deducciones",    cop(ded["otras"])])
    ded_rows.append(["TOTAL DEDUCCIONES", cop(ded.get("total_deducciones", 0))])
    story.append(make_table(ded_rows))

    story.append(Spacer(1, 0.2*cm))
    neto_t = Table([["NETO A PAGAR", cop(neto)]], colWidths=[10*cm, 4.5*cm])
    neto_t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), NAV),
        ("TEXTCOLOR",  (0, 0), (-1, -1), colors.white),
        ("FONTNAME",   (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 12),
        ("ALIGN",      (1, 0), (1, -1), "RIGHT"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING",   (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
    ]))
    story.append(neto_t)

    story.append(Paragraph("CARGA PATRONAL", H2))
    pat_rows = [["Concepto", "Valor"],
        ["Salud empleador (8.5%)",    cop(pat.get("salud_empleador", 0))],
        ["Pensión empleador (12%)",   cop(pat.get("pension", 0))],
        ["ARL",                       cop(pat.get("arl", 0))],
        ["SENA (2%)",                 cop(pat.get("sena", 0))],
        ["ICBF (3%)",                 cop(pat.get("icbf", 0))],
        ["Caja de Compensación (4%)", cop(pat.get("caja_compensacion", 0))],
        ["TOTAL CARGA PATRONAL",      cop(pat.get("total", 0))]]
    story.append(make_table(pat_rows))

    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(
        f"Generado por TaxOps SaaS · {date.today().strftime('%d/%m/%Y')} · "
        "CST Colombia · Art.114-1 ET · Ley 1607/2012 · Ley 2466/2025",
        FOOT,
    ))

    doc.build(story)
    buf.seek(0)
    safe_nombre = emp_nombre.replace(" ", "_")
    filename = f"colilla_{safe_nombre}_{periodo}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})


# ═══════════════════════════════════════════════════════════════════════════════
# EMAIL — envío de colillas vía Resend
# ═══════════════════════════════════════════════════════════════════════════════

@router.post("/periodos/{periodo}/email")
async def enviar_colillas_email(periodo: str, body: dict, user: dict = Depends(get_current_user)) -> dict:
    """Envía colillas PDF por email usando Resend API."""
    resend_key = os.getenv("RESEND_API_KEY", "")
    if not resend_key:
        raise HTTPException(status_code=503, detail="RESEND_API_KEY no configurada")

    from_email = os.getenv("RESEND_FROM_EMAIL", "noreply@taxops.co")
    from sqlalchemy import text
    import base64
    import urllib.request

    get_db = _get_db()
    with get_db() as db:
        per = db.execute(
            text("SELECT id FROM nomina_periodos WHERE org_id = :org AND periodo = :p LIMIT 1"),
            {"org": user["org_id"], "p": periodo},
        ).mappings().fetchone()
        if not per:
            raise HTTPException(status_code=404, detail="Período no encontrado. Calcula primero.")

        detalles = db.execute(
            text("""
                SELECT d.resultado, d.empleado_id, e.nombre, e.email
                FROM nomina_detalle d
                JOIN empleados e ON e.id = d.empleado_id
                WHERE d.periodo_id = :pid AND e.email != ''
                ORDER BY e.nombre
            """),
            {"pid": per["id"]},
        ).mappings().fetchall()

    # Filtrar empleados con email
    empleado_ids = body.get("empleado_ids")
    sent = []
    skipped = []
    errors = []

    for det in detalles:
        emp_id = str(det["empleado_id"])
        if empleado_ids and emp_id not in empleado_ids:
            continue
        if not det["email"]:
            skipped.append(det["nombre"])
            continue

        # Generar PDF en memoria
        try:
            pdf_buf = await _generar_colilla_pdf(
                resultado=det["resultado"] if isinstance(det["resultado"], dict) else json.loads(det["resultado"]),
                nombre=det["nombre"],
                periodo=periodo,
            )
            pdf_b64 = base64.b64encode(pdf_buf).decode()
        except Exception as exc:
            errors.append(f"{det['nombre']}: error PDF — {exc}")
            continue

        # Enviar vía Resend
        payload = json.dumps({
            "from": from_email,
            "to":   [det["email"]],
            "subject": f"Colilla de nómina {periodo} — TaxOps",
            "html": (
                f"<p>Hola <strong>{det['nombre']}</strong>,</p>"
                f"<p>Adjuntamos tu colilla de nómina del período <strong>{periodo}</strong>.</p>"
                "<p>Este es un mensaje automático generado por <strong>TaxOps</strong>.</p>"
            ),
            "attachments": [{
                "filename": f"colilla_{periodo}.pdf",
                "content":  pdf_b64,
            }],
        }).encode()

        req = urllib.request.Request(
            "https://api.resend.com/emails",
            data=payload,
            headers={"Authorization": f"Bearer {resend_key}", "Content-Type": "application/json"},
            method="POST",
        )
        try:
            with urllib.request.urlopen(req, timeout=15) as resp:
                resp.read()
            sent.append(det["nombre"])
        except Exception as exc:
            errors.append(f"{det['nombre']}: error Resend — {exc}")

    return {"enviados": sent, "sin_email": skipped, "errores": errors}


async def _generar_colilla_pdf(resultado: dict, nombre: str, periodo: str) -> bytes:
    """Helper que genera bytes de PDF para una colilla (reutilizado por email)."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    NAV = colors.HexColor("#1A3A5C")
    ORG = colors.HexColor("#E05519")
    LGT = colors.HexColor("#EEF2F7")
    styles = getSampleStyleSheet()
    H1 = ParagraphStyle("h1", parent=styles["Normal"], fontSize=14, textColor=colors.white,
                        backColor=NAV, alignment=TA_CENTER, fontName="Helvetica-Bold", leading=20, spaceAfter=2)
    H2 = ParagraphStyle("h2", parent=styles["Normal"], fontSize=9, textColor=colors.white,
                        backColor=ORG, spaceAfter=1, spaceBefore=6, fontName="Helvetica-Bold", leading=13)
    FOOT = ParagraphStyle("foot", parent=styles["Normal"], fontSize=7, textColor=colors.grey, alignment=TA_CENTER)

    def make_table(rows):
        t = Table(rows, colWidths=[10*cm, 4.5*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), NAV),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LGT]),
            ("GRID",       (0, 0), (-1, -1), 0.25, colors.HexColor("#DDDDDD")),
            ("ALIGN",      (1, 0), (1, -1), "RIGHT"),
            ("LEFTPADDING",  (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING",   (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
        ]))
        return t

    def cop(v): return f"${float(v or 0):,.0f}"

    dev  = resultado.get("devengado", {})
    ded  = resultado.get("deducciones", {})
    pat  = resultado.get("carga_patronal", {})
    neto = resultado.get("neto_pagar", 0)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=1.8*cm, leftMargin=1.8*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    story = []
    story.append(Paragraph(f"TaxOps · Colilla de Nómina · {periodo}", H1))
    story.append(Paragraph(f"Empleado: {nombre}", styles["Normal"]))
    story.append(Spacer(1, 0.3*cm))

    dev_rows = [["Concepto", "Valor"],
        ["Salario proporcional", cop(dev.get("salario_proporcional", 0))]]
    if dev.get("aplica_aux_transporte"):
        dev_rows.append(["Auxilio de transporte", cop(dev.get("auxilio_transporte", 0))])
    he = dev.get("horas_extras", {})
    for lbl, key in [("HED +25%","hed"),("HEN +80%","hen"),("RN +35%","rn"),
                     ("HOD +80%","hod"),("HEDD +105%","hedd"),("HEND +155%","hend"),("RND +115%","rnd")]:
        if he.get(key, 0):
            dev_rows.append([f"  · {lbl}", cop(he[key])])
    dev_rows.append(["TOTAL DEVENGADO", cop(dev.get("total_devengado", 0))])
    story.append(Paragraph("DEVENGADO", H2))
    story.append(make_table(dev_rows))

    ded_rows = [["Concepto", "Valor"],
        ["IBC – Base Seg. Social",    cop(ded.get("ibc", 0))],
        ["Salud empleado (4%)",        cop(ded.get("salud_empleado", 0))],
        ["Pensión empleado (4%)",      cop(ded.get("pension_empleado", 0))]]
    if ded.get("fsp"):             ded_rows.append(["FSP",              cop(ded["fsp"])])
    if ded.get("retencion_fuente"):ded_rows.append(["Retención fuente", cop(ded["retencion_fuente"])])
    ded_rows.append(["TOTAL DEDUCCIONES", cop(ded.get("total_deducciones", 0))])
    story.append(Paragraph("DEDUCCIONES", H2))
    story.append(make_table(ded_rows))

    neto_t = Table([["NETO A PAGAR", cop(neto)]], colWidths=[10*cm, 4.5*cm])
    neto_t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), NAV), ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
        ("FONTNAME",   (0, 0), (-1, -1), "Helvetica-Bold"), ("FONTSIZE",  (0, 0), (-1, -1), 12),
        ("ALIGN",      (1, 0), (1, -1), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING",  (0, 0), (-1, -1), 6), ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
    ]))
    story.append(Spacer(1, 0.2*cm))
    story.append(neto_t)
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(
        f"Generado por TaxOps SaaS · {date.today().strftime('%d/%m/%Y')}",
        FOOT,
    ))
    doc.build(story)
    return buf.getvalue()
